from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


def export_results_to_excel(exam):
    """
    Exports the results of answer sheets for an exam to an Excel file.
    
    Args:
        exam: Instance of the Exam model
    
    Returns:
        BytesIO: Buffer containing the generated Excel file
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    # Configure header
    ws['A1'] = f"AVALIAÇÃO: {exam.subject_name}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    # Column headers
    headers = ['CODE', 'ITENS CORRETOS', 'ITENS INCORRETOS', 'PERCENTAGE']
    ws.append([])  # Blank line
    ws.append(headers)

    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Fetch all student answer sheets for this exam
    student_sheets = exam.student_answer_sheets.all().order_by('-accuracy_percentage')

    # Add data
    for sheet in student_sheets:
        ws.append([
            sheet.sheet_code,
            sheet.correct_items,
            sheet.incorrect_items,
            f"{float(sheet.accuracy_percentage):.2f}%"
        ])

    # Adjust column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 15

    # Add statistics at the end
    if student_sheets.exists():
        ws.append([])  # Blank line
        ws.append(['STATISTICS'])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

        # Calculate statistics
        total_students = student_sheets.count()
        avg_correct = sum([s.correct_items for s in student_sheets]) / total_students if total_students > 0 else 0
        avg_percentage = sum(
            [float(s.accuracy_percentage) for s in student_sheets]) / total_students if total_students > 0 else 0

        ws.append(['Total students:', total_students])
        ws.append(['Average correct items:', f"{avg_correct:.2f}"])
        ws.append(['Average percentage:', f"{avg_percentage:.2f}%"])

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer


def export_detailed_results_to_excel(exam):
    """
    Exports detailed results of answer sheets for an exam to Excel,
    including the answers for each question.
    
    Args:
        exam: Instance of the Exam model
    
    Returns:
        BytesIO: Buffer containing the generated Excel file
    """
    wb = Workbook()

    # Sheet 1: Summary
    ws_summary = wb.active
    ws_summary.title = "Resultados Resumidos"

    ws_summary['A1'] = f"EXAM: {exam.subject_name}"
    ws_summary['A1'].font = Font(bold=True, size=14)
    ws_summary.merge_cells('A1:E1')

    headers_summary = ['Código', 'Itens Corretos', 'Itens Incorretos', 'Percentual']
    ws_summary.append([])
    ws_summary.append(headers_summary)

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")

    for col in range(1, len(headers_summary) + 1):
        cell = ws_summary.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    student_sheets = exam.student_answer_sheets.all().order_by('-accuracy_percentage')

    for sheet in student_sheets:
        ws_summary.append([
            sheet.sheet_code,
            sheet.correct_items,
            sheet.incorrect_items,
            f"{float(sheet.accuracy_percentage):.2f}%"
        ])

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    # Sheet 2: Detailed Answers
    ws_details = wb.create_sheet(title="Respostas Detalhadas")

    ws_details['A1'] = f"Detalhes de respostas - {exam.subject_name}"
    ws_details['A1'].font = Font(bold=True, size=14)

    # Create dynamic headers based on number of questions
    headers_details = ['CODE']
    for q in range(1, exam.num_questions + 1):
        headers_details.append(f'Q{q}')
    headers_details.extend(['CORRETO', 'INCORRETO', '%'])

    ws_details.append([])
    ws_details.append(headers_details)

    for col in range(1, len(headers_details) + 1):
        cell = ws_details.cell(row=3, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Add detailed data
    try:
        correct_answer_sheet = exam.correct_answer_sheet
        correct_answers = correct_answer_sheet.answers

        for sheet in student_sheets:
            row_data = [
                sheet.sheet_code,
            ]

            # Add answers for each question
            if sheet.student_answers:
                for q in range(1, exam.num_questions + 1):
                    student_answer = sheet.student_answers.get(str(q), '-')
                    correct_answer = correct_answers.get(str(q), '-')

                    # Mark if correct or incorrect
                    if student_answer == correct_answer:
                        row_data.append(f"{student_answer} ✓")
                    else:
                        row_data.append(f"{student_answer} ✗")
            else:
                row_data.extend(['-'] * exam.num_questions)

            row_data.extend([
                sheet.correct_items,
                sheet.incorrect_items,
                f"{float(sheet.accuracy_percentage):.2f}%"
            ])

            ws_details.append(row_data)

        # Adjust column widths
        ws_details.column_dimensions['A'].width = 20
        ws_details.column_dimensions['B'].width = 30
        for col_idx in range(3, 3 + exam.num_questions):
            ws_details.column_dimensions[chr(64 + col_idx)].width = 8

    except Exception as e:
        ws_details.append(['Error generating details:', str(e)])

    # Save to buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer
